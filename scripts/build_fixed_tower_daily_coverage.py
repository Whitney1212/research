#!/usr/bin/env python
"""Build daily fixed-tower coverage tables from the Level0 processing workbook.

The script reads the current Level0 processing record, merges registered
coverage intervals for MT/CVT EC/AP/MET, and writes a natural-year daily
calendar plus stream summaries and gap segments.
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path("E:/Dataset_Level0/数据存入-处理记录.xlsx")
DEFAULT_OUTPUT_DIR = Path("E:/Dataset_Level1/FixedTowerCoverage")
VALID_STATUS = {"1", "RAW", "OGA", "raw", "oga"}
FULL_DAY_HOURS = 23.5


@dataclass(frozen=True)
class StreamRule:
    stream: str
    sheet: str
    start_col: int
    end_col: int
    status_col: int


RULES = [
    StreamRule("MT_EC", "MT_EC", 0, 1, 2),
    StreamRule("MT_AP", "MT_AP", 0, 1, 2),
    StreamRule("MT_MET", "MT_MET", 0, 1, 2),
    # CVT_EC uses the second 6-column block in the current workbook.
    StreamRule("CVT_EC", "CVT_EC", 6, 7, 8),
    StreamRule("CVT_AP", "CVT_AP", 0, 1, 2),
    StreamRule("CVT_MET", "CVT_MET", 0, 1, 2),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build fixed-tower daily coverage tables for one year."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--year", type=int, default=2025)
    return parser.parse_args()


def parse_datetime(value, is_end: bool) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        base = datetime.combine(value, time.min)
        return base.replace(hour=23, minute=59, second=59) if is_end else base

    text = str(value).strip()
    if not text:
        return None

    digits = re.sub(r"\D", "", text)
    if len(digits) >= 12:
        return datetime.strptime(digits[:12], "%Y%m%d%H%M")
    if len(digits) == 8:
        base = datetime.strptime(digits, "%Y%m%d")
        return base.replace(hour=23, minute=59, second=59) if is_end else base

    match = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", text)
    if match:
        y, m, d = map(int, match.groups())
        base = datetime(y, m, d)
        return base.replace(hour=23, minute=59, second=59) if is_end else base

    return None


def merge_intervals(intervals: Iterable[tuple[datetime, datetime]]):
    ordered = sorted(intervals)
    merged: list[list[datetime]] = []
    for start, end in ordered:
        if end < start:
            continue
        if not merged or start > merged[-1][1]:
            merged.append([start, end])
        elif end > merged[-1][1]:
            merged[-1][1] = end
    return [(start, end) for start, end in merged]


def load_coverage(workbook: Path):
    if not workbook.exists():
        raise FileNotFoundError(f"Missing workbook: {workbook}")

    wb = load_workbook(workbook, data_only=True, read_only=True)
    coverage = {}
    raw_counts = {}

    for rule in RULES:
        ws = wb[rule.sheet]
        intervals = []
        valid_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= max(rule.start_col, rule.end_col, rule.status_col):
                continue
            status = str(row[rule.status_col]).strip()
            if status not in VALID_STATUS:
                continue
            start = parse_datetime(row[rule.start_col], is_end=False)
            end = parse_datetime(row[rule.end_col], is_end=True)
            if start is None or end is None or end < start:
                continue
            intervals.append((start, end))
            valid_rows += 1
        coverage[rule.stream] = merge_intervals(intervals)
        raw_counts[rule.stream] = valid_rows

    return coverage, raw_counts


def day_bounds(day: date):
    start = datetime.combine(day, time.min)
    return start, start + timedelta(days=1)


def overlap_hours(day: date, intervals: list[tuple[datetime, datetime]]) -> float:
    day_start, day_end = day_bounds(day)
    seconds = 0.0
    for start, end in intervals:
        overlap_start = max(start, day_start)
        overlap_end = min(end, day_end)
        if overlap_end > overlap_start:
            seconds += (overlap_end - overlap_start).total_seconds()
    return round(seconds / 3600.0, 4)


def iter_year_days(year: int):
    day = date(year, 1, 1)
    end = date(year + 1, 1, 1)
    while day < end:
        yield day
        day += timedelta(days=1)


def bool_text(value: bool) -> str:
    return "TRUE" if value else "FALSE"


def build_daily_rows(year: int, coverage):
    rows = []
    streams = [rule.stream for rule in RULES]
    for day in iter_year_days(year):
        row = {
            "date": day.isoformat(),
            "year": year,
            "doy": day.timetuple().tm_yday,
        }
        full_flags = {}
        any_flags = {}
        for stream in streams:
            hours = overlap_hours(day, coverage[stream])
            any_flag = hours > 0
            full_flag = hours >= FULL_DAY_HOURS
            row[f"{stream}_hours"] = f"{hours:.4f}"
            row[f"{stream}_any"] = bool_text(any_flag)
            row[f"{stream}_full_day"] = bool_text(full_flag)
            any_flags[stream] = any_flag
            full_flags[stream] = full_flag

        pair_defs = {
            "AP_PAIR": ["MT_AP", "CVT_AP"],
            "EC_PAIR": ["MT_EC", "CVT_EC"],
            "MET_PAIR": ["MT_MET", "CVT_MET"],
            "ALL_SIX": streams,
        }
        for name, members in pair_defs.items():
            row[f"{name}_any"] = bool_text(all(any_flags[x] for x in members))
            row[f"{name}_full_day"] = bool_text(all(full_flags[x] for x in members))

        row["n_streams_any"] = sum(any_flags.values())
        row["n_streams_full_day"] = sum(full_flags.values())
        if row["ALL_SIX_full_day"] == "TRUE":
            row["coverage_class"] = "all_six_full_day"
            row["gap_note"] = ""
        elif row["AP_PAIR_full_day"] == "TRUE":
            row["coverage_class"] = "ap_primary_full_day_covariates_incomplete"
            row["gap_note"] = "EC/MET covariates incomplete or maintenance/registered gap"
        elif row["AP_PAIR_any"] == "TRUE":
            row["coverage_class"] = "ap_primary_partial_day"
            row["gap_note"] = "AP pair has partial-day coverage"
        else:
            row["coverage_class"] = "primary_ap_missing"
            row["gap_note"] = "AP pair unavailable"
        rows.append(row)
    return rows


def summarise_daily(rows):
    keys = [rule.stream for rule in RULES] + ["AP_PAIR", "EC_PAIR", "MET_PAIR", "ALL_SIX"]
    summary = []
    for key in keys:
        any_col = f"{key}_any"
        full_col = f"{key}_full_day"
        if any_col not in rows[0]:
            any_col = f"{key}_any"
        total = len(rows)
        any_days = sum(row[any_col] == "TRUE" for row in rows)
        full_days = sum(row[full_col] == "TRUE" for row in rows)
        partial_days = any_days - full_days
        no_days = total - any_days
        summary.append({
            "stream_or_pair": key,
            "total_days": total,
            "any_coverage_days": any_days,
            "full_day_coverage_days": full_days,
            "partial_coverage_days": partial_days,
            "no_coverage_days": no_days,
        })
    return summary


def build_gap_segments(rows):
    keys = [rule.stream for rule in RULES] + ["AP_PAIR", "EC_PAIR", "MET_PAIR", "ALL_SIX"]
    segments = []
    for key in keys:
        full_col = f"{key}_full_day"
        any_col = f"{key}_any"
        current = None
        for row in rows:
            is_gap = row[full_col] != "TRUE"
            gap_kind = "no_coverage" if row[any_col] != "TRUE" else "partial_coverage"
            day = date.fromisoformat(row["date"])
            if not is_gap:
                if current is not None:
                    segments.append(current)
                    current = None
                continue
            if current is None:
                current = {
                    "stream_or_pair": key,
                    "gap_start": row["date"],
                    "gap_end": row["date"],
                    "gap_days": 1,
                    "gap_kind": gap_kind,
                    "gap_note": "registered gap / maintenance candidate",
                }
            else:
                prev_end = date.fromisoformat(current["gap_end"])
                if day == prev_end + timedelta(days=1) and current["gap_kind"] == gap_kind:
                    current["gap_end"] = row["date"]
                    current["gap_days"] += 1
                else:
                    segments.append(current)
                    current = {
                        "stream_or_pair": key,
                        "gap_start": row["date"],
                        "gap_end": row["date"],
                        "gap_days": 1,
                        "gap_kind": gap_kind,
                        "gap_note": "registered gap / maintenance candidate",
                    }
        if current is not None:
            segments.append(current)
    return segments


def write_csv(path: Path, rows: list[dict]):
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        raise ValueError(f"No rows to write: {path}")
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_run_notes(path: Path, args, coverage, raw_counts, rows, summary, gaps):
    path.parent.mkdir(parents=True, exist_ok=True)
    full_all_six = next(x for x in summary if x["stream_or_pair"] == "ALL_SIX")
    ap_pair = next(x for x in summary if x["stream_or_pair"] == "AP_PAIR")
    with path.open("w", encoding="utf-8") as handle:
        handle.write("Fixed tower daily coverage build\n")
        handle.write(f"Workbook: {args.workbook}\n")
        handle.write(f"Year: {args.year}\n")
        handle.write(f"Output directory: {args.output_dir}\n")
        handle.write("\nRegistered rows used by stream:\n")
        for stream, count in raw_counts.items():
            handle.write(f"- {stream}: {count}\n")
        handle.write("\nKey daily counts:\n")
        handle.write(f"- AP_PAIR full-day days: {ap_pair['full_day_coverage_days']}\n")
        handle.write(f"- ALL_SIX full-day days: {full_all_six['full_day_coverage_days']}\n")
        handle.write(f"- Total gap segments: {len(gaps)}\n")
        handle.write("\nInterpretation notes:\n")
        handle.write("- TRUE/FALSE coverage is derived from registered Level0 intervals.\n")
        handle.write("- Gap segments are registered gaps or maintenance candidates; the workbook does not encode a finer reason.\n")
        handle.write("- Repair gaps should be excluded from no-peak control days unless instrument logs prove usable data.\n")


def main():
    args = parse_args()
    coverage, raw_counts = load_coverage(args.workbook)
    rows = build_daily_rows(args.year, coverage)
    summary = summarise_daily(rows)
    gaps = build_gap_segments(rows)

    prefix = f"fixed_tower_coverage_{args.year}"
    write_csv(args.output_dir / f"{prefix}_daily.csv", rows)
    write_csv(args.output_dir / f"{prefix}_summary.csv", summary)
    write_csv(args.output_dir / f"{prefix}_gap_segments.csv", gaps)
    write_run_notes(args.output_dir / f"{prefix}_run_notes.txt", args, coverage, raw_counts, rows, summary, gaps)

    ap_pair = next(x for x in summary if x["stream_or_pair"] == "AP_PAIR")
    all_six = next(x for x in summary if x["stream_or_pair"] == "ALL_SIX")
    print(f"Wrote outputs to {args.output_dir}")
    print(f"AP_PAIR full-day days: {ap_pair['full_day_coverage_days']}")
    print(f"ALL_SIX full-day days: {all_six['full_day_coverage_days']}")
    print(f"Gap segments: {len(gaps)}")


if __name__ == "__main__":
    main()
